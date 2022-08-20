from selenium import webdriver
import urllib.request
import time
import re
import openpyxl
import pyautogui as pg
import os
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


#url_first = 'https://www.youtube.com/results?search_query=%E3%83%97%E3%83%AD%E3%82%B0%E3%83%A9%E3%83%9F%E3%83%B3%E3%82%B0&sp=EgIQAg%253D%253D'
#twitterログイン
driver = webdriver.Chrome(executable_path="C:\\Users\\tisk0\\PycharmProjects\\chromedriver_win32\\chromedriver.exe")
driver.maximize_window()
driver.implicitly_wait(3)
driver.get('https://twitter.com/login?lang=ja')
id_element = driver.find_element_by_css_selector('#react-root > div > div > div.css-1dbjc4n.r-13qz1uu.r-417010 > main > div > div > div.css-1dbjc4n.r-13qz1uu > form > div > div:nth-child(6) > label > div > div.css-1dbjc4n.r-18u37iz.r-16y2uox.r-1wbh5a2.r-1wzrnnt.r-1udh08x.r-xd6kpl.r-1pn2ns4.r-ttdzmv > div > input')
id_element.send_keys('yomi_creating')
password_element = driver.find_element_by_css_selector('#react-root > div > div > div.css-1dbjc4n.r-13qz1uu.r-417010 > main > div > div > div.css-1dbjc4n.r-13qz1uu > form > div > div:nth-child(7) > label > div > div.css-1dbjc4n.r-18u37iz.r-16y2uox.r-1wbh5a2.r-1wzrnnt.r-1udh08x.r-xd6kpl.r-1pn2ns4.r-ttdzmv > div > input')
password_element.send_keys('pcstrong777')
button_element = driver.find_element_by_css_selector('#react-root > div > div > div.css-1dbjc4n.r-13qz1uu.r-417010 > main > div > div > div.css-1dbjc4n.r-13qz1uu > form > div > div:nth-child(8) > div > div')
button_element.click()
time.sleep(4)
driver.execute_script("window.open()")
driver.switch_to.window(driver.window_handles[1])

#アップロード順にする
url_original_former = 'https://www.youtube.com/results?search_query='
url_original_lattter = '&sp=CAISAhAC'
url_list = ['きりぬき']
for url_first in url_list:
    done_channel_name = []
    done_channel_twitter_url = []
    done_channel_number = []
    try:
        #ループ
        driver.get(url_original_former + url_first + url_original_lattter)
        #スクロール
        time.sleep(3)
        for n in range(100):
            #pg.hotkey('end', 'ctrl')
            webdriver.ActionChains(driver).key_down(Keys.END).key_down(Keys.CONTROL).perform()
            time.sleep(1)

        time.sleep(3)
        channel_url_elements = driver.find_elements_by_id('main-link')
        channel_url = []
        for channel_url_element in channel_url_elements:
            channel_url.append(channel_url_element.get_attribute("href"))

        channel_name_elements = driver.find_elements_by_class_name('style-scope ytd-channel-name')
        channel_name = []
        for channel_name_element in channel_name_elements:
            name = channel_name_element.get_attribute("textContent").splitlines()[2]
            channel_name.append(name[4:])

        channel_number_elements = driver.find_elements_by_id('subscribers')
        channel_numbers = []
        for channel_number_element in channel_number_elements:
            channel_number_kari = channel_number_element.text[10:]
            channel_numbers.append(channel_number_kari)

        #print(channel_number_elements[0].text)

        channel_name = channel_name[:-2]

        #print(len(channel_name))
        #print(len(channel_url))
        #print(len(channel_numbers))
        total = len(channel_numbers)
        i = 0
        twi_count = 1
        for channel_number in channel_numbers:
            #print(channel_number)
            try:
                if '万' in channel_number:
                    #print('含む')
                    channel_number = channel_number[:-2]
                    #print(channel_number)
                    if float(channel_number) < 30:
                        #投稿1週間前以内かcheck
                        url = channel_url[i]
                        driver.get(url + '/videos')
                        if len(driver.find_elements_by_css_selector('#metadata-line > span:nth-child(2)')) > 0:
                            video_date_element = driver.find_element_by_css_selector('#metadata-line > span:nth-child(2)')
                            video_date = video_date_element.text
                            #video_date = video_date[:3]
                            #print(video_date)
                            if '時間' in video_date or '日' in video_date or '1 週' in video_date or '分' in video_date:
                                driver.get(url + '/about')
                                #twitter確認

                                #SNSが複数ある場合の対処

                                if len(driver.find_elements_by_css_selector('#link-list-container > a')) > 0:
                                    for sns in driver.find_elements_by_css_selector('#link-list-container > a'):
                                        #twitter_element = driver.find_element_by_css_selector('#link-list-container > a')
                                        twitter_url = sns.get_attribute("href")
                                        if 'twitter' in twitter_url:
                                            driver.get(twitter_url)
                                            twitter_url = driver.current_url
                                            time.sleep(1)
                                            if i == 0:
                                                driver.refresh()
                                            #DM確認
                                            if len(driver.find_elements_by_css_selector('#react-root > div > div > div.css-1dbjc4n.r-18u37iz.r-13qz1uu.r-417010 > main > div > div > div > div.css-1dbjc4n.r-14lw9ot.r-jxzhtn.r-1ljd8xs.r-13l2t4g.r-1phboty.r-1jgb5lz.r-11wrixw.r-61z16t.r-1ye8kvj.r-13qz1uu.r-184en5c > div > div:nth-child(2) > div > div > div:nth-child(1) > div > div.css-1dbjc4n.r-obd0qt.r-18u37iz.r-1w6e6rj.r-1wtj0ep > div > div:nth-child(2) > div')) > 0:
                                                #存在する
                                                #print(channel_name[i])
                                                #print('DMあり')
                                                done_channel_name.append(channel_name[i])
                                                done_channel_twitter_url.append(twitter_url)
                                                done_channel_number.append(channel_numbers[i])
                                                print(twi_count)
                                                twi_count = twi_count + 1

                else:
                    url = channel_url[i]
                    driver.get(url + '/videos')
                    if len(driver.find_elements_by_css_selector('#metadata-line > span:nth-child(2)')) > 0:
                        video_date_element = driver.find_element_by_css_selector('#metadata-line > span:nth-child(2)')
                        video_date = video_date_element.text
                        #video_date = video_date[:3]
                        #print(video_date)
                        if '時間' in video_date or '日' in video_date or '1 週' in video_date or '分' in video_date:

                            driver.get(url + '/about')
                            # twitter確認
                            if len(driver.find_elements_by_css_selector('#link-list-container > a')) > 0:
                                for sns in driver.find_elements_by_css_selector('#link-list-container > a'):
                                    #twitter_element = driver.find_element_by_css_selector('#link-list-container > a')
                                    twitter_url = sns.get_attribute("href")
                                    if 'twitter' in twitter_url:
                                        driver.get(twitter_url)
                                        twitter_url = driver.current_url
                                        time.sleep(1)
                                        if i == 0:
                                            driver.refresh()
                                        # DM確認
                                        if len(driver.find_elements_by_css_selector('#react-root > div > div > div.css-1dbjc4n.r-18u37iz.r-13qz1uu.r-417010 > main > div > div > div > div.css-1dbjc4n.r-14lw9ot.r-jxzhtn.r-1ljd8xs.r-13l2t4g.r-1phboty.r-1jgb5lz.r-11wrixw.r-61z16t.r-1ye8kvj.r-13qz1uu.r-184en5c > div > div:nth-child(2) > div > div > div:nth-child(1) > div > div.css-1dbjc4n.r-obd0qt.r-18u37iz.r-1w6e6rj.r-1wtj0ep > div > div:nth-child(2) > div')) > 0:
                                            # 存在する
                                            #print(channel_name[i])
                                            #print('DMあり')
                                            done_channel_name.append(channel_name[i])
                                            done_channel_twitter_url.append(twitter_url)
                                            done_channel_number.append(channel_numbers[i])
                                            print(twi_count)
                                            twi_count = twi_count + 1
            except:
                print('エラー発生')
            i = i + 1
            print(str(i+1) + '/' + str(total))
    except:
        print('エラー発生')

    # Excel作成
    wb = openpyxl.Workbook()
    ws = wb["Sheet"]
    for i in range(0, len(done_channel_name)):
        # A列にリストを書き込み
        ws.cell(i + 1, 1, value=done_channel_name[i])

        # B列にリストを書き込み
        ws.cell(i + 1, 2, value=done_channel_twitter_url[i])

        ws.cell(i + 1, 3, value=done_channel_number[i])

    # ファイル名連番にする
    # wb.save("完成.xlsx")
    k = 1
    while True:
        if os.path.isfile('完成' + str(k) + '.xlsx'):
            k = k + 1
        else:
            wb.save('完成' + str(k) + '.xlsx')
            break


